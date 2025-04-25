"""
Google Search Automation Script
- Performs searches for given terms
- Extracts top results
- Saves to formatted Excel file
- Uses stealth techniques to avoid detection
"""

import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium_stealth import stealth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urlparse  # For URL parsing

# ========== BROWSER SETUP ==========
# Configure ChromeDriver path and options
service = Service("C:/Users/huria/My git projects/my-first-python-code/chromedriver.exe")  # Path to your ChromeDriver

# Chrome options to customize browser behavior
options = webdriver.ChromeOptions()
options.add_argument("user-data-dir=C:/Users/huria/AppData/Local/Google/Chrome/User Data")  # Use your Chrome profile
options.add_argument("--start-maximized")  # Start browser maximized

# Launch Chrome browser
driver = webdriver.Chrome(service=service, options=options)

# ========== STEALTH CONFIGURATION ==========
# Make selenium look more like a real browser to avoid detection
stealth(driver,
        languages=["en-US", "en"],  # Browser languages
        vendor="Google Inc.",       # Browser vendor
        platform="Win32",          # Operating system
        webgl_vendor="Intel Inc.", # Graphics card vendor
        renderer="Intel Iris OpenGL Engine",  # Graphics renderer
        fix_hairline=True,         # Fix for subtle rendering differences
)

# ========== EXCEL FILE SETUP ==========
# Create a new Excel workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Google Search Results"  # Rename the sheet

# Define styles for Excel formatting
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue header
header_font = Font(color="FFFFFF", bold=True, size=12)  # White, bold header text
cell_font = Font(size=11)  # Regular cell font
border = Border(left=Side(style='thin'),  # Thin borders for all cells
               right=Side(style='thin'), 
               top=Side(style='thin'), 
               bottom=Side(style='thin'))
center_aligned = Alignment(horizontal='center')  # Center alignment
left_aligned = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Left alignment with text wrap

# Set up column headers
headers = ["Search Term", "Rank", "Result Title", "URL", "Domain", "Timestamp"]
for col_num, header in enumerate(headers, 1):  # Start from column 1
    cell = ws.cell(row=1, column=col_num, value=header)  # Write header
    cell.fill = header_fill  # Apply blue background
    cell.font = header_font  # Apply white bold font
    cell.border = border     # Add borders
    cell.alignment = center_aligned  # Center align

# Set column widths (in characters)
column_widths = [25, 8, 50, 60, 20, 20]  # Custom widths for each column
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width  # Apply widths

# ========== SEARCH TERMS ==========
search_terms = [
    "Python tutorial", 
    "Selenium automation", 
    "GitHub basics", 
    "VS Code tips"
]  # Terms to search for

# ========== HELPER FUNCTIONS ==========
def get_search_results():
    """
    Extracts search results from Google results page
    Returns list of tuples: (rank, title, url)
    """
    results = []
    try:
        # Wait until search results are loaded (looking for the main search div)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#search"))
        )
        
        # Try multiple selector patterns to find results (Google changes these frequently)
        selectors = [
            ("h3", "./.."),  # Pattern 1: h3 heading with parent being the link
            ("div.g a", None),  # Pattern 2: Classic Google result link
            ("a[href][ping]", None),  # Pattern 3: Link with ping attribute
            ("div.rc a", None)  # Pattern 4: Alternative container
        ]
        
        found_results = False
        
        # Try each selector pattern until we find results
        for selector, parent_xpath in selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                if not elements:
                    continue  # Skip if no elements found with this selector
                    
                # Process the first 3 results
                for i, element in enumerate(elements[:3], 1):
                    try:
                        # Get the actual link element
                        if parent_xpath:
                            link = element.find_element(By.XPATH, parent_xpath)
                        else:
                            link = element
                        
                        # Get title - either from h3 or link text
                        title = element.text if element.tag_name == "h3" else link.text
                        url = link.get_attribute("href")
                        
                        # Skip Google's own links (like "Images", "Maps", etc.)
                        if url and "google.com" not in url:
                            results.append((i, title if title else "No title", url))
                            found_results = True
                    except:
                        continue  # Skip if any error occurs with this element
                
                if found_results:
                    break  # Stop if we found results with this selector
                    
            except:
                continue  # Skip to next selector if error
        
        # If no results found with standard selectors, try a broader approach
        if not found_results:
            try:
                links = driver.find_elements(By.CSS_SELECTOR, "#search a")
                for i, link in enumerate(links[:3], 1):
                    url = link.get_attribute("href")
                    if url and "google.com" not in url:
                        results.append((i, link.text or "No title", url))
            except:
                pass  # Give up if we can't find anything
        
        return results if results else [(1, "No results found", "N/A")]
        
    except Exception as e:
        print(f"Error in get_search_results: {str(e)}")
        return [(1, "Error retrieving results", "N/A")]

# ========== MAIN SEARCH LOOP ==========
for row, term in enumerate(search_terms, start=2):  # Start from row 2 (below headers)
    try:
        print(f"\nSearching for: {term}")
        
        # Step 1: Load Google homepage
        driver.get("https://www.google.com")
        time.sleep(1)  # Small delay to allow page load
        
        # Step 2: Accept cookies if the popup appears
        try:
            accept_button = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Accept all')]"))
            )
            accept_button.click()
            print("Accepted cookies")
            time.sleep(1)
        except:
            pass  # Continue if no cookie popup appears
        
        # Step 3: Find search box and enter query
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "q"))
        )
        search_box.clear()  # Clear any existing text
        
        # Type the search term character by character (more human-like)
        for char in term:
            search_box.send_keys(char)
            time.sleep(0.1 + (0.2 if char.isupper() else 0))  # Type slower for uppercase
        
        # Submit search
        search_box.send_keys(Keys.RETURN)
        print("Search submitted")
        
        # Step 4: Wait for results to load
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#search"))
        )
        time.sleep(1)  # Additional wait for stability
        
        # Step 5: Extract results
        results = get_search_results()
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Step 6: Write results to Excel
        for i, (rank, title, url) in enumerate(results):
            current_row = row + i  # Calculate current row (main term + results)
            
            # Extract domain from URL (e.g., "www.example.com" → "example.com")
            domain = urlparse(url).netloc.replace("www.", "") if url not in ["N/A", "Not found"] else "N/A"
            
            # Write data to Excel cells
            ws.cell(row=current_row, column=1, value=term if i == 0 else "").font = cell_font
            ws.cell(row=current_row, column=2, value=rank).font = cell_font
            ws.cell(row=current_row, column=3, value=title).font = cell_font
            ws.cell(row=current_row, column=4, value=url).font = Font(color="0066CC", underline="single")  # Blue underlined URL
            ws.cell(row=current_row, column=5, value=domain).font = cell_font
            ws.cell(row=current_row, column=6, value=timestamp).font = cell_font
            
            # Apply borders and alignment to all cells
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=col).alignment = left_aligned if col in [1,3,4] else center_aligned
            
            print(f"  Found result {rank}: {title[:50]}... | {domain}")  # Print first 50 chars of title
        
        time.sleep(2)  # Polite delay between searches
        
    except Exception as e:
        print(f"Error processing '{term}': {str(e)}")
        # Mark error in Excel
        ws.cell(row=row, column=1, value=term).font = Font(color="FF0000")  # Red text
        ws.cell(row=row, column=3, value=f"Error: {str(e)}").font = Font(color="FF0000")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border

# ========== FINAL EXCEL FORMATTING ==========
ws.freeze_panes = "A2"  # Freeze header row so it stays visible when scrolling

# Add summary information at the bottom
summary_row = len(search_terms) * 3 + 3  # Calculate row for summary
ws.cell(row=summary_row, column=1, value="Search Summary").font = Font(bold=True, size=12)
ws.cell(row=summary_row, column=2, value=f"Total searches: {len(search_terms)}").font = Font(bold=True)
ws.cell(row=summary_row+1, column=2, value=f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

# ========== SAVE EXCEL FILE ==========
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"Google_Search_Results_{timestamp}.xlsx"
wb.save(filename)
print(f"\n✅ Report successfully generated: {filename}")

# ========== CLEAN UP ==========
driver.quit()  # Close the browser
input("Press Enter to exit...")  # Keep console open until user presses Enter